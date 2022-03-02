import os
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from datetime import datetime
from django.http import HttpResponse, HttpResponseBadRequest, HttpResponseNotFound
from django.shortcuts import get_object_or_404
from rest_framework.views import APIView
from rest_framework.response import Response
from rir_data.models.indicator import IndicatorValue
from rir_data.models.instance import Instance

from openpyxl.workbook import Workbook
from django.conf import settings


class DownloadMasterData(APIView):
    """
    Download master data as spreadsheet
    """

    def get(self, request, slug, date):
        instance = get_object_or_404(
            Instance, slug=slug
        )
        try:
            date = datetime.strptime(date, "%Y-%m-%d").date()
        except ValueError:
            return HttpResponseBadRequest('Date format is not correct')
        today_date = datetime.today().date().strftime("%Y-%m-%d")

        folder = os.path.join(
            settings.MEDIA_ROOT, 'download-data', request.user.username, today_date)
        if not os.path.exists(folder):
            os.makedirs(folder)
        _file = os.path.join(folder, date.strftime("%Y-%m-%d") + '.xlsx')

        # ----------------------------------------------
        # ----------------- SAVING DATA ----------------
        # ----------------------------------------------
        wb = Workbook()
        del wb['Sheet']

        # filter indicators
        indicators = instance.indicators.order_by('name')
        if request.GET.get('indicator_id_in', None):
            indicators = indicators.filter(id__in=request.GET.get('indicator_id_in', None).split(','))

        geometries = instance.geometries()

        level_trees = {}
        for instance_level in instance.geometry_instance_levels:
            level_trees[instance_level.level.name] = instance_level.get_level_tree()

        # fix this should be per instance levels
        for instance_level in instance.geometry_instance_levels:
            # get top geometry
            trees = instance_level.get_level_tree()
            geometry_parent = geometries.filter(geometry_level__name=trees[len(trees) - 1]).first()

            if not geometry_parent:
                continue

            # create headers
            header = [f'{instance_level.level.name} Name', f'{instance_level.level.name} Code']

            # we create the indicators based on the reporting level of indicator
            # is the level is the child of the instance_level
            level_indicators = []
            for idx, indicator in enumerate(indicators):
                level_tree = level_trees[indicator.geometry_reporting_level.name]
                if instance_level.level.name in level_tree:
                    level_indicators.append(indicator)

            # check if there is indicators
            if not level_indicators:
                continue

            # create new sheet
            sheet = wb.create_sheet(instance_level.level.name)

            def insert_sheet(row, column, value, color=None):
                sheet.cell(row=row + 1, column=column + 1).value = value
                if color:
                    try:
                        sheet.cell(row=row + 1, column=column + 1).fill = PatternFill(
                            "solid", fgColor=color.replace('#', ''))
                    except ValueError:
                        pass

            # get lines per geometry
            geometry_code_rows = ['']
            for row, geometry in enumerate(geometries.filter(
                    geometry_level=instance_level.level).order_by('name'), 1):
                insert_sheet(row, 0, geometry.name)
                insert_sheet(row, 1, geometry.identifier)
                geometry_code_rows.append(geometry.identifier)

            for indicator in level_indicators:
                values = indicator.values(
                    geometry=geometry_parent, geometry_level=instance_level.level, date_data=date
                )
                if not values:
                    continue
                header.append(indicator.name)
                header.append(f'{indicator.name} value')
                for value in values:
                    row = geometry_code_rows.index(value['geometry_code'])
                    indicator_text_col = header.index(indicator.name)
                    indicator_value_col = header.index(f'{indicator.name} value')
                    insert_sheet(row, indicator_text_col, value['scenario_text'], value['background_color'])
                    insert_sheet(row, indicator_value_col, value['scenario_value'], value['background_color'])

            # safe header to excel
            for idx, _header in enumerate(header):
                insert_sheet(0, idx, _header)
                sheet.column_dimensions[get_column_letter(idx + 1)].width = 20

        # save output reports to excel file
        wb.save(filename=_file)
        # ----------------------------------------------

        if os.path.exists(_file):
            with open(_file, "rb") as excel:
                data = excel.read()

            response = HttpResponse(data, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename={date.strftime("%Y-%m-%d")}_Master_Data.xlsx'
            os.remove(_file)
            return response
        else:
            return HttpResponseNotFound('The file is not found')


class DownloadMasterDataCheck(APIView):
    """
    Download master data as spreadsheet
    """

    def get(self, request, slug, date):
        instance = get_object_or_404(
            Instance, slug=slug
        )
        try:
            date = datetime.strptime(date, "%Y-%m-%d").date()
        except ValueError:
            return HttpResponseBadRequest('Date format is not correct')
        indicators = instance.indicators
        if request.GET.get('indicator_id_in', None):
            indicators = indicators.filter(id__in=request.GET.get('indicator_id_in', None).split(','))

        return Response({
            'count': IndicatorValue.objects.filter(
                indicator__in=indicators).filter(
                date__lte=date
            ).count()
        })
